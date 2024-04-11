import openpyxl
import pickle 
import os
import json

default_config = {
  'input_col_num': 29,
  'search_col_num': 2,
  'begin_name_row_num': 2,
  'ending_name_row_num': 96,
  'name_col': 'AC',
  'file_excel_name':'c1dhl_1.xlsx'
}

path_config = 'config.pkl'
a= os.path.isfile(path_config)

def loadConfig():
  print('Setting')
  with open('config.pkl', 'rb') as file: 
    # A new file will be created 
    data = pickle.load(file) 
    return data

def saveConfig(config):
  print('Setting')
  with open('config.pkl', 'wb') as file: 
    # A new file will be created 
    pickle.dump(config, file)

if(a==False):
  saveConfig(default_config)
## opening the previously created xlsx file using 'load_workbook()' method
xlsx = openpyxl.load_workbook(default_config['file_excel_name'])
## getting the sheet to active
sheet = xlsx.active
# Open a file and use dump() 
with open('config.pkl', 'rb') as file: 
    # A new file will be created 
    # pickle.dump(students, file) 
    data = pickle.load(file)
    print(data)

print('1. Nhập số phút gọi. ')
print('2. Setting. ')
print('3. Thoát. ')
print(' ------------------------------------- ')

def checkSelectName(founded_names, select):
  for key, value in founded_names.items():
    if(int(select) == int(key)):
      return 1
  return 0

def checkName(name_csi):
  founded_name_data = {}
  bol_check_emp = 0

  for number in range(default_config['begin_name_row_num'], default_config['ending_name_row_num']):
    tag = sheet.cell(row = number, column = default_config['search_col_num'])
    stt = sheet.cell(row = number, column = 1)
    if(name_csi.upper() in tag.value):
      founded_name_data[stt.value] = tag.value
      id_name = stt.value
      bol_check_emp = 1
    elif(number == default_config['ending_name_row_num']):
      if(bol_check_emp == 0):
        print('not found anyone in list.')
  for key, value in founded_name_data.items():
    print(str(key) + '. '+str(value))
  if(len(founded_name_data) == 1):
    return founded_name_data
  print('---------------')
  print('nhập id tên muốn thao tác:')
  select = input()
  if(checkSelectName(founded_name_data, select) == 1):
    print('Đã chọn: '+str(founded_name_data[int(select)]))
    select_obj = {key: str(founded_name_data[int(select)])}
    return select_obj
  return 0

  

select_option = input()
if(select_option == '1') :
  i=1
  while(i > 0):
    print('---------------------------------------')
    print('')
    print('Name: ')
    pre_name = input()
    if(pre_name == 'Q' or pre_name == 'q'):
      print('')
      print('Thoát khỏi chương trình, bye bye... ')
      break
    try:
        # Loading the json string using json.loads() method
        s = json.dumps(checkName(pre_name))
        present_name = json.loads(s)
        print('hello')
        print(present_name)
    except json.decoder.JSONDecodeError as e:
        # Catching the exception and printing the error message
        print("Invalid json string:", e)
    print('time (minutes): ')
    call_time = input()
    if(call_time == 'Q' or call_time =='q'):
      i=-1
      print('')
      print('Thoát khỏi chương trình, bye bye... ')
      break

    for key, value in present_name.items():
      if(sheet.cell(row = int(key)+1, column = default_config['input_col_num']).value is not None):
        number_time_previous = sheet.cell(row = int(key)+1, column = default_config['input_col_num']).value
        print(number_time_previous)
        sheet[default_config['name_col']+str(int(key)+1)] = str(number_time_previous) +'+'+ str(call_time)
        xlsx.save(default_config['file_excel_name'])
        print('OK.')
        break
      else:
        sheet[default_config['name_col']+str(int(key)+1)] = '='+ call_time
        xlsx.save(default_config['file_excel_name'])
        print('OK.')
        break


    ## getting the reference of the cells which we want to get the data from
    # names = sheet[row_[0]:row_[1]]
    # for number in range(default_config['begin_name_row_num'], default_config['ending_name_row_num']):
    #   tag = sheet.cell(row = number, column = default_config['search_col_num'])
    #   stt = sheet.cell(row = number, column = 1)
    #   if(pre_name.upper() in tag.value):
    #     if(sheet.cell(row = number, column = default_config['input_col_num']).value is not None):
    #       number_time_previous = sheet.cell(row = number, column = default_config['input_col_num']).value
    #       sheet[default_config['name_col']+str(number)] = str(number_time_previous) +'+'+ str(call_time)
    #       xlsx.save(default_config['file_excel_name'])
    #       print('OK.')
    #       break
    #     else:
    #       sheet['AC'+str(number)] = '='+ call_time
    #       xlsx.save(default_config['file_excel_name'])
    #       print('OK.')
    #       break
    #   elif(number == 95):
    #     print('Không tìm thấy tên trong danh sách.')
    #     break

if(select_option =='2'):
  print('-------------------------------')
  print('1. Thay đổi cột nhập dữ liệu.')
  print('2. Thay đổi cột tìm kiếm.')
  print('3. Thay đổi khoảng tên nhập liệu. (tìm kiếm tên từ cột 2 đến cột 96).')
  print('Enter your choise: ')
  choise = input()
  if(choise == '1'):
    print('Nhập số cột nhập dữ liệu.')
    choice_1 = input()
    if(choice_1 > 0 and choice_1 < 111 and choice_1.isnumeric):
      default_config
      

if(select_option =='3'):
  print('Exited.')
else:
  print('')
  print ('đã thoát!')


